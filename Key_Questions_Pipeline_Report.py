#!/usr/bin/env python
# coding: utf-8

# In[1]:


print("Welcome")
# Pipeline is Open opportunities that are Stage 1 & beyond (exclude stage 0)
# to get all Open opportunities you just need to also include stage 0. This metric is rarely used though.

#import requests
import datetime
import yaml #pyyaml
import numpy as np
from datetime import datetime
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
#import dataframe_image as dfi #https://randomds.com/2021/12/23/visualize-and-save-full-pandas-dataframes-as-images/
# looks useful: https://geo.rocks/post/python-to-exe/ from https://stackoverflow.com/questions/55741607/is-it-possible-to-generate-an-executable-exe-of-a-jupyter-notebook
# Merging in Pandas: https://stackoverflow.com/questions/53645882/pandas-merging-101

# Hold credentials in .yaml?
# Use python to hit call some code in an excel file? MAybe use batch as an intermediary?

# Pandas Pivot Table Filter:
# https://stackoverflow.com/questions/43235930/adding-filter-to-pandas-pivot-table
# Pivot table basics: https://builtin.com/data-science/pandas-pivot-tables



# In[2]:


# Imports username & security_token from setup.yaml
# get security token from... https://tigertext.lightning.force.com/lightning/settings/personal/ResetApiToken/home
with open("setup.yaml") as stream:
    try:
        #print(yaml.safe_load(stream))
        values = yaml.safe_load(stream)
        #print(values)
    except yaml.YAMLError as exc:
        pass
        print(exc)
#print(values)


# In[3]:


import pandas as pd
import csv
import requests
from io import StringIO
from simple_salesforce import Salesforce
from getpass import getpass

# Input Salesforce credentials:
sf = Salesforce(
    username=values["username"], 
    password= getpass(),# gets input from Command prompt, hides what you type 
    security_token=values["security_token"])
print("Please wait while I extract that data for you")
# Basic report URL structure:
orgParams = 'https://tigertext.my.salesforce.com/' # you can see this in your Salesforce URL
exportParams = '?isdtp=p1&export=1&enc=UTF-8&xf=csv'

# Downloading the report:
#reportId = '00O5x0000083qfOEAQ' # You find this in the URL of the report in question between "Report/" and "/view"
#reportId = '00O5x000008GgNREA0'
#reportId = '00O5x00000848trEAA'
reportId = '00O5x000008GgycEAC'
reportUrl = orgParams + reportId + exportParams
reportReq = requests.get(reportUrl, headers=sf.headers, cookies={'sid': sf.session_id})
reportData = reportReq.content.decode('utf-8')
# If your report comes out empty, make sure you don't have any f(x) lines included. Idk maybe remove groups too never tried those
df = pd.read_csv(StringIO(reportData))


# In[4]:


# Basic report URL structure:
orgParams = 'https://tigertext.my.salesforce.com/' # you can see this in your Salesforce URL
exportParams = '?isdtp=p1&export=1&enc=UTF-8&xf=csv'

# Downloading the report:
reportId = '00O5x000007ULYLEA4'
reportUrl = orgParams + reportId + exportParams
reportReq = requests.get(reportUrl, headers=sf.headers, cookies={'sid': sf.session_id})
reportData = reportReq.content.decode('utf-8')
# If your report comes out empty, make sure you don't have any f(x) lines included. Idk maybe remove groups too never tried those
userDf = pd.read_csv(StringIO(reportData))
userDf.drop(userDf.tail(5).index, inplace = True) #Drop saleforce info junk lines at bottom
userDf


# In[5]:


df.drop(df.tail(5).index, inplace = True) #Drop saleforce info junk lines at bottom
# Convert 'Close Date (2)' column to datetime object
df['Close Date']= pd.to_datetime(df['Close Date'])
df


# In[6]:


today = datetime.today()
date1 = today.strftime('%y%m%d')
print(date1)


# In[7]:


import pipreqsnb 
pipreqsnb


# In[8]:


# Different cut of the data
#date2023 = '2022-12-31'
#date2024 = '2022-12-31'
#pre2023 = df[(df["Close Date"] >= date1) & (df["Close Date"] <= date2)]
#df['Close Date (2)']= pd.to_datetime(df['Close Date (2)'])
#df['Close Date']= pd.to_datetime(df['Close Date'])
df['Stage 1 Date Stamp']= pd.to_datetime(df['Stage 1 Date Stamp'])
df['ARR'] = df['MRR + Est MRR'] * 12
df.loc[ df["Stage"] == "6. Contracting", 'Stage'] = "Contracting" # new line to map the old contracting stage
df['Stage Bucket'] = df['Stage']
df.loc[ df["Stage"] == "Active Buying", 'Stage Bucket'] = "Open" #not needed and seemingly innaccurate?
df.loc[ df["Stage"] == "Active Selling", 'Stage Bucket'] = "Open"
df.loc[ df["Stage"] == "Contracting", 'Stage Bucket'] = "Open"
df.loc[ df["Stage"] == "Discovery", 'Stage Bucket'] = "Open"
df.loc[ df["Stage"] == "Qualified", 'Stage Bucket'] = "Open"
# Apparently these are considered Open??? Salesforce you confuse me you undocumented fuck
df.loc[ df["Stage"] == "Prospect Deciding", 'Stage Bucket'] = "Open"
df.loc[ df["Stage"] == "Sales Complete", 'Stage Bucket'] = "Open"
df.loc[ df["Stage"] == "6. Contracting", 'Stage Bucket'] = "Open"

# Stage 1 = Qualified
# Stage 2 = Discovery
# Stage 3 = Active Selling
# Stage 4 = Active Buying
# Stage 5 = Contracting


awl = df[(df['Stage Bucket'] == "Open")] #wrong total
openTotal = df[(df['Closed'] == 0) & (df['Won'] == 0) & (df['Team'] == 'Enterprise') & (df['Team'] != 'SMB Sales')] #correct total?
pivot = pd.pivot_table(openTotal, values = "ARR", index="Stage", aggfunc=np.sum, margins = True) 
pivot


# In[9]:


TotalKeyQuestions = df[(df["Close Date"].dt.year == 2023)& (df['Team'] == 'Enterprise') & (df['Team'] != 'SMB')& (df['Team'] != 'SMB Sales')]
pivot = pd.pivot_table(TotalKeyQuestions, values = "ARR", index="Stage Bucket", aggfunc=np.sum, margins = True) 
pivot
#TotalKeyQuestions.to_csv("KeyQuestionsData.csv")
#the two below won't quite add to this total due to opps without a Stage 1 date stamp. 


# In[10]:


# Key Questions & Analysis pipeline figures
#here for excel sheet, copy of this line below for it's own pivot.
pipe2023 = df[(df["Close Date"].dt.year == 2023) &  (df["Stage 1 Date Stamp"].dt.year == 2023)& (df['Team'] == 'Enterprise') & (df['Team'] != 'SMB')& (df['Team'] != 'SMB Sales')]
#pre2023 = df2.
#pivot = pd.pivot_table(pre2023, values = "MRR + Est MRR", index="Stage", aggfunc=np.sum, margins = True) 
pre2023 = df[(df["Close Date"].dt.year == 2023) & (df["Stage 1 Date Stamp"].dt.year < 2023)& (df['Team'] == 'Enterprise')& (df['Team'] != 'SMB')& (df['Team'] != 'SMB Sales')]# & (df['Closed'] == 0) & (df['Won'] == 0)]
pivot = pd.pivot_table(pre2023, values = "ARR", index="Stage Bucket", aggfunc=np.sum, margins = True) 
#pre2023.to_csv("KeyQuestionsPre2023.csv")

# Load the workbook
book = load_workbook('DONOTTOUCH.xlsx')
            
#writer = pd.ExcelWriter(date1 + ' Key Questions.xlsx', engine='openpyxl') 
#writer.book = book

# Then write the dataframe to the sheet
#pre2023.to_excel(writer, index=False, sheet_name='Pre 2023 Pipeline')
#pre2023.to_excel(writer, index=False, sheet_name='Pre2023Pipeline')



# Save the changes
#writer.save()
#writer.close()
import shutil

# copy original file
shutil.copy('DONOTTOUCH.xlsx', date1 + ' Key Questions.xlsx')
book = load_workbook('DONOTTOUCH.xlsx')

sheet = book['Pre2023Pipeline']
rows = dataframe_to_rows(pre2023, index=False, header=True)
for i, row in enumerate(rows, 1):
    for j, value in enumerate(row, 1):
        sheet.cell(row=i, column=j, value=value)
        
sheet = book['2023 Pipeline']
rows = dataframe_to_rows(pipe2023, index=False, header=True)
for i, row in enumerate(rows, 1):
    for j, value in enumerate(row, 1):
        sheet.cell(row=i, column=j, value=value)

# Save the workbook
book.save(date1 + ' Key Questions.xlsx')
#book.save('Key Questions_copy.xlsx')
pivot


# In[11]:


pipe2023 = df[(df["Close Date"].dt.year == 2023) &  (df["Stage 1 Date Stamp"].dt.year == 2023)& (df['Team'] == 'Enterprise') & (df['Team'] != 'SMB')& (df['Team'] != 'SMB Sales')]
pivot = pd.pivot_table(pipe2023, values = "ARR", index="Stage Bucket", aggfunc=np.sum, margins = True) 
#pipe2023.to_csv("KeyQuestionsPipe2023.csv")
pivot


# In[12]:


#pipeWTF = df[(df["Close Date"].dt.year == 2023) &  (df["Stage 1 Date Stamp"].dt.year != 2023)]
#pivot = pd.pivot_table(pipeWTF, values = "ARR", index="Stage 1 Date Stamp", aggfunc=np.sum, margins = True) 
#pivot
pipe2023 = df[(df["Close Date"].dt.year == 2023) &  (df["Stage 1 Date Stamp"].dt.year == 2023) & (df["Stage Bucket"] == "Open")& (df['Team'] == 'Enterprise') & (df['Team'] != 'SMB')& (df['Team'] != 'SMB Sales')]
pivot = pd.pivot_table(pipe2023, values = "ARR", index="Stage", aggfunc=np.sum, margins = True) 
pivot


# In[13]:


pre2023 = df[(df["Close Date"].dt.year == 2023) &  (df["Stage 1 Date Stamp"].dt.year < 2023) & (df["Stage Bucket"] == "Open")& (df['Team'] == 'Enterprise') & (df['Team'] != 'SMB')& (df['Team'] != 'SMB Sales')]
pivot = pd.pivot_table(pre2023, values = "ARR", index="Stage", aggfunc=np.sum, margins = True) 
pivot


# In[14]:


listErrorEntPS = pipe2023[pipe2023['Stage'] == 'Prospect Deciding']
listErrorEntPS["Team"]

